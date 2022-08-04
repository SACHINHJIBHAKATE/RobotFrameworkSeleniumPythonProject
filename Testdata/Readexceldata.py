import openpyxl

# APPROACH 1 (basic code)

# This is to print the number of rows within the sheet in excel
workbook = openpyxl.load_workbook("../Testdata/data.xlsx")
# sheet = workbook["Sheet1"]
# print(sheet.max_row)

# This is to print the value in cell with row 1 & column 1
# format = sheet.cell(row, column)
# cell = sheet.cell(1, 1)
# print(cell.value)

# APPROACH 2 (pass the sheetname, row number & column number as arguments from the method)

# def fetch_number_of_rows_in_a_sheet (sheet_name):
#     sheet = workbook[sheet_name]
#     print(sheet.max_row)

# def fetch_cell_data (sheet_name, row, column):
#     sheet = workbook[sheet_name]
#     cell=sheet.cell(row,column)
#     print(cell.value)

# fetch_number_of_rows_in_a_sheet ("Sheet1")
# fetch_cell_data ("Sheet1", 1, 1)

# APPROACH 3 (pass the sheetname, row number & column number as arguments from the method & return values from the actual methods)

def fetch_number_of_rows_in_a_sheet (sheet_name):
    sheet = workbook[sheet_name]
    return sheet.max_row

def fetch_cell_data (sheet_name, row, column):
    sheet = workbook[sheet_name]
    cell=sheet.cell(int(row),int(column))
    return cell.value

# print(fetch_number_of_rows_in_a_sheet ("Sheet1"))
# print(fetch_cell_data ("Sheet1", 1, 1))






